import json
import openpyxl
import os

base_url = os.path.dirname(os.path.abspath(__file__))
project_dir = os.path.split(base_url)[0]

class HandleExcel(object):
    """操作Excel的类"""

    def __init__(self, file_path=os.path.join(project_dir, "files\\testcase.xlsx"), sheet_name=None, sheet=0):
        if os.path.exists(file_path):
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
        else:
            raise FileNotFoundError("excel:%s not found" % file_path)
        if sheet_name:
            self.sheet_value = self.wb[sheet_name]
        else:
            self.sheet_value = self.wb[self.wb.sheetnames[sheet]]
        self.file_name = file_path

    def get_sheet_value(self, sheet_name=None, sheet=0):
        if sheet_name:
            sheet_value = self.wb[sheet_name]
        else:
            sheet_value = self.wb[self.wb.sheetnames[sheet]]
        return sheet_value

    def get_cell_value(self, row=None, col=None, coor=None):
        """
        获取单元格内容
        :param row:单元格行数
        :param col:单元格列数
        :param coor:单元格坐标，比如A5
        :return:返回单元格内容
        """
        if row and col:
            return self.sheet_value.cell(row=row, column=col).value
        if coor:
            return self.sheet_value[coor].value

    def get_rows(self):
        """获取Excel的行数"""
        try:
            return self.sheet_value.max_row
        except AttributeError:
            print("请检查case路径，sheet等")
            return

    def get_row_values(self, row):
        """获取某一行数据"""
        if row == 0:
            raise IndexError("row请输入大于0的数字")
        row_list = list()

        rows = self.sheet_value[row]
        for i in rows:
            data = i.value
            if data is not None and '\n' in data:
                data = data.replace('\n', '')
            row_list.append(data)
        return row_list

    def get_column_values(self, col='A'):
        """获取某一列的数据"""
        columns_list = []
        sheet_values = self.sheet_value[col]
        for i in sheet_values:
            columns_list.append(i.value)
        return columns_list

    def write_excel(self, row, col, text):
        if isinstance(text, dict):
            text = json.dumps(text, ensure_ascii=False, indent=2)
        try:
            self.sheet_value.cell(row, col, text)
            self.wb.save(self.file_name)
        except Exception as e:
            print("%s 保存失败" % self.file_name)
            print(str(e))

    def get_all_data(self):
        """获取excel所有内容，放到一个list中"""
        data_list = []
        len = self.get_rows()
        for j in range(2, len+1):
            data = self.get_row_values(j)
            if data[2] == "yes":
                data_list.append(data)

        return data_list

    def get_row_number(self, case_id):
        """根据case id获取行号"""
        cases = self.get_column_values("A")
        n = len(cases)
        for i in range(n):
            if case_id == cases[i]:
                return i+1
        #
        # sheet_values = self.sheet_value['A']
        # return sheet_values

    def depend_data(self, data):
        case_ids = data.split(";")
        case_rows = list()
        for id in case_ids:
            row_number = self.get_row_number(id)
            values = self.get_row_values(row_number)
            case_rows.append(values)
        return case_rows


if __name__ == '__main__':
    wb = HandleExcel(r'..\files\testcase.xlsx', 0)
    # a = wb.get_row_number("diy_004")
    # b = wb.get_sheet_value("Sheet1")
    #
    # c = wb.depend_data("diy_commit_001<diy_commit_002")
    a = wb.get_all_data()
    for i in a:
        print(i)
    n = wb.get_rows()





